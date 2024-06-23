import os
from openpyxl import Workbook

dirname, filename = os.path.split(os.path.abspath(__file__))

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Add some data to the worksheet
ws['A1'] = 10
ws['A2'] = 20
ws['A3'] = 30

# Set a formula in cell B1 to sum the values of cells A1, A2, and A3
ws['B1'] = '=SUM(A1:A3)'

# Set a formula in cell B2 to calculate the average of cells A1, A2, and A3
ws['B2'] = '=AVERAGE(A1:A3)'

# Set a formula in cell B3 to calculate the product of cells A1, A2, and A3
ws['B3'] = '=PRODUCT(A1:A3)'

# Set a formula in cell B4 to get the maximum value from cells A1, A2, and A3
ws['B4'] = '=MAX(A1:A3)'

# Set a formula in cell B5 to get the minimum value from cells A1, A2, and A3
ws['B5'] = '=MIN(A1:A3)'

# Save the workbook
wb.save(f'{dirname}/formulas_example.xlsx')
