import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

dirname, filename = os.path.split(os.path.abspath(__file__))

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set the font style
font = Font(name="Calibri", size=14, bold=True, italic=True, color="FF0000")

# Set the fill pattern
fill = PatternFill(fill_type="solid", fgColor="FFFF00")

# Set the border style
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Set the alignment
alignment = Alignment(horizontal="center", vertical="center")

# Apply styles to a cell
cell = ws["A1"]
cell.value = "Styled Cell"
cell.font = font
cell.fill = fill
cell.border = thin_border
cell.alignment = alignment


# Loop through a range of cells and apply styles
for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=3):
    for cell in row:
        cell.value = "Styled"
        cell.font = font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = alignment


# Save the workbook
wb.save(f"{dirname}/styled_example.xlsx")
