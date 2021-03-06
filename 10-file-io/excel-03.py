from openpyxl import load_workbook

# WRITE
wb = load_workbook(filename="students.xlsx")
sheet = wb.active

# WRITE SINGLE CELL
sheet['A1'] = "nameeeee"
sheet.cell(row=1, column=3).value = "gpaaaaa"

# WRITE MULTIPLE CELL
rows = (
    ('name', 'sex', 'gpa'),
    ('Olan', 'M', 1.9),
    ('Game', 'M', 3.9),
    ('Jenny', 'F', 1.9),
    ('Lisa', 'F', 3.9),
    ('Jim', 'M', 1.9),
    ('GG', 'F', 3.9)
)

# for row in rows :
#     sheet.append(row)

for i, row in enumerate(rows):
    for j, cell in enumerate(row) :
        sheet.cell(row=i + 1, column=j + 1).value = cell

wb.save("students.xlsx")