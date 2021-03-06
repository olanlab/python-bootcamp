from openpyxl import load_workbook

# READ
wb = load_workbook(filename = "students.xlsx")
sheet = wb.active

# READ SINGLE CELL
print(sheet['A1'].value)
print(sheet.cell(row=1, column=1).value)

# READ MULTIPLE CELLS
cells = sheet['A1' : 'C4']
for row in cells :
    for cell in row :
        print(cell.value, end = " ")
    print()

# READ ITERATE
for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column) :
    for cell in row :
        print(cell.value, end = " ")
    print()